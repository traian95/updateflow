"""
Extrage din bare.xlsx: categorie, cod model, lungime, decor (INOX/NEGRU), pret EUR.
Genereaza JSON si CSV potrivite pentru import in Supabase (inserare in tabel).

Exemplu:
  python scripts/extract_bare.py
  python scripts/extract_bare.py --input bare.xlsx --out-json data/bare_preturi.json
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

# Coloane Excel (1-based): A=cod model, D=lungime, E+=preturi per decor
DECOR_START_COL = 5


def _clean_model(raw) -> str | None:
    if raw is None:
        return None
    s = str(raw).replace("\n", " ").replace("\r", " ").strip()
    if not s:
        return None
    return re.sub(r"\s+", " ", s)


def _parse_lungime_cm(lungime_raw) -> int | None:
    if lungime_raw is None:
        return None
    text = str(lungime_raw).strip()
    m = re.search(r"(\d+)\s*cm", text, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def _parse_pret(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val != val:  # NaN
            return None
        return float(val)
    s = str(val).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _read_decors_from_header(ws, row_idx: int) -> list[tuple[int, str]]:
    """Returneaza [(col_index_1based, nume_decor), ...] din prima linie de date."""
    out: list[tuple[int, str]] = []
    for col in range(DECOR_START_COL, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col).value
        if cell is None:
            continue
        name = str(cell).strip()
        if not name:
            continue
        if "pret" in name.lower() and "lista" in name.lower():
            continue
        out.append((col, name))
    return out


def extract_rows(xlsx_path: Path) -> tuple[list[dict], str, str]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    categorie_raw = ws.cell(row=1, column=1).value
    categorie = str(categorie_raw).strip() if categorie_raw else "BARA"

    decors = _read_decors_from_header(ws, 1)
    if not decors:
        raise ValueError(
            f"Nu s-au gasit coloane decor in {xlsx_path} (rand 1, coloana E+)."
        )

    nota_pret = ws.cell(row=2, column=DECOR_START_COL).value
    sursa_pret = str(nota_pret).strip() if nota_pret else None

    rows_out: list[dict] = []
    current_model: str | None = None

    for r in range(3, ws.max_row + 1):
        model_cell = _clean_model(ws.cell(row=r, column=1).value)
        if model_cell:
            current_model = model_cell

        lungime_raw = ws.cell(row=r, column=4).value
        lungime_text = str(lungime_raw).strip() if lungime_raw is not None else ""
        if not lungime_text or not current_model:
            continue

        lungime_cm = _parse_lungime_cm(lungime_raw)

        for col, decor in decors:
            pret = _parse_pret(ws.cell(row=r, column=col).value)
            if pret is None:
                continue
            rows_out.append(
                {
                    "categorie": categorie,
                    "cod_model": current_model,
                    "lungime_text": lungime_text,
                    "lungime_cm": lungime_cm,
                    "decor": decor,
                    "pret_eur": pret,
                    "moneda": "EUR",
                    "tva_inclus": False,
                    "sursa_pret": sursa_pret,
                }
            )

    return rows_out, categorie, sursa_pret or ""


def main() -> int:
    root = Path(__file__).resolve().parent.parent
    p = argparse.ArgumentParser(description="Extrage preturi bare din Excel pentru Supabase.")
    p.add_argument(
        "--input",
        type=Path,
        default=root / "bare.xlsx",
        help="Cale catre bare.xlsx (implicit: radacina proiectului).",
    )
    p.add_argument(
        "--out-json",
        type=Path,
        default=root / "bare_preturi.json",
        help="Fisier JSON (array de obiecte).",
    )
    p.add_argument(
        "--out-csv",
        type=Path,
        default=root / "bare_preturi.csv",
        help="Fisier CSV (acelasi continut ca JSON).",
    )
    p.add_argument("--no-csv", action="store_true", help="Nu scrie CSV.")
    args = p.parse_args()

    if not args.input.is_file():
        print(f"Fisier inexistent: {args.input}", file=sys.stderr)
        return 1

    try:
        records, categorie, _ = extract_rows(args.input)
    except Exception as e:
        print(f"Eroare la citire: {e}", file=sys.stderr)
        return 1

    if not records:
        print("Nu s-au extras inregistrari.", file=sys.stderr)
        return 1

    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(
        json.dumps(records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    if not args.no_csv:
        args.out_csv.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = list(records[0].keys())
        with args.out_csv.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(records)

    print(
        f"Extrase {len(records)} linii (model x lungime x decor). "
        f"Categorie: {categorie}. JSON: {args.out_json}"
        + ("" if args.no_csv else f", CSV: {args.out_csv}")
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
